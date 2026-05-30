"""Generate webapp/public/data/rgint_meta.json.

Combines the RGINT index (id/name/uf/biome), centroids + planimetric area
from the IBGE RG2017 shapefile, and municipality counts from the 2017
composition spreadsheet.
"""
from __future__ import annotations

import json

import geopandas as gpd
import openpyxl

from common import INDEX_JSON, MUNICIPIOS_XLSX, SHAPEFILE, ensure_out, load_index

# SIRGAS 2000 / Brazil Polyconic — equal-area-ish projection for hectares.
EQUAL_AREA_EPSG = 5880


def municipio_counts() -> dict[str, int]:
    wb = openpyxl.load_workbook(MUNICIPIOS_XLSX, read_only=True)
    ws = wb[wb.sheetnames[0]]
    header = [str(c) if c is not None else "" for c in next(ws.iter_rows(values_only=True))]
    rg_col = header.index("cod_rgint")
    counts: dict[str, int] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        rgint = str(row[rg_col]).strip()
        if rgint and rgint != "None":
            counts[rgint] = counts.get(rgint, 0) + 1
    return counts


def main() -> None:
    index = load_index()
    by_id = {item["rgint"]: item for item in index}

    gdf = gpd.read_file(SHAPEFILE)
    gdf["rgint"] = gdf["rgint"].astype(str)
    projected = gdf.to_crs(epsg=EQUAL_AREA_EPSG)
    centroids = projected.geometry.centroid.to_crs(epsg=4326)
    areas_ha = projected.geometry.area / 10_000.0

    counts = municipio_counts()

    meta = []
    for idx, row in gdf.iterrows():
        rgint = row["rgint"]
        info = by_id.get(rgint, {})
        meta.append(
            {
                "id": rgint,
                "nome": info.get("nome", row["nome_rgint"]),
                "uf": info.get("uf", ""),
                "bioma_principal": info.get("biome", ""),
                "area_ha": round(float(areas_ha.iloc[idx]), 2),
                "n_municipios": counts.get(rgint, 0),
                "lat_centroide": round(float(centroids.iloc[idx].y), 5),
                "lon_centroide": round(float(centroids.iloc[idx].x), 5),
            }
        )

    meta.sort(key=lambda m: m["id"])
    out = ensure_out() / "rgint_meta.json"
    out.write_text(json.dumps(meta, ensure_ascii=False, indent=0), encoding="utf-8")
    print(f"wrote {out} ({len(meta)} regions)")


if __name__ == "__main__":
    assert INDEX_JSON.exists(), "rgint_index.json missing"
    main()
